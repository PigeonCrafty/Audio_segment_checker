'''
Hello, this script is designed to help you check the accuracy of audio segments against the runsheet. 
To run this script, please follow these steps: 

1) Ensure that you have pre-installed [Python] and the following packages via pip: 
- openpyxl
- SpeechRecognition

2) Place all of your local audio segments and the runsheet in the same folder.

3) In the runsheet, create a separate sheet with the audio file names in the first column and the corresponding texts in the second column. Make sure that this sheet is in ACTIVATE status.

When you run the script and are prompted, please enter the [language code] for the audio segments (e.g. 'en-US' for English, 'zh-CN' for Chinese, 'ja-JP' for Japanese, etc.). 
You will also need to provide the [path] for both the audio segments and the runsheet. 

Complete language code list for Azure Speech Service: 
https://learn.microsoft.com/en-us/azure/cognitive-services/speech-service/language-support?tabs=stt
'''

import os
import openpyxl
import speech_recognition as sr
from difflib import SequenceMatcher

def transcribe_audio(filename, language):
    r = sr.Recognizer()
    # azure_key = os.environ.get("AZURE_KEY")
    # azure_region = os.environ.get("AZURE_REGION") 
    azure_key = "yourAzureSpeechServiceKey"  # Please replace with your own Azure key
    azure_region = "yourAzureRegion"                         # Please replace with your own Azure region, eg: "eastus"
    with sr.AudioFile(filename) as source:
        audio_data = r.record(source)
        response = r.recognize_azure(audio_data, key=azure_key, location=azure_region, language=language)
        return response[0]

def compare_transcript_to_runsheet(transcript, runsheet):
    return SequenceMatcher(None, transcript, runsheet).ratio()

def read_runsheet(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    runsheet = {row[0]: row[1] for row in sheet.iter_rows(min_row=1, values_only=True)}
    return runsheet

def main():
    print("""
Hello, this script is designed to help you check the accuracy of audio segments against the runsheet. 
To run this script, please follow these steps: 

1) Ensure that you have pre-installed [Python] and the following packages via pip: 
- openpyxl
- SpeechRecognition
2) Place all of your local audio segments and the runsheet in the same folder.
3) In the runsheet, create a separate sheet with the audio file names in the first column and the corresponding texts in the second column. Make sure that this sheet is in ACTIVATE status.

When you run the script and are prompted, please enter the [language code] for the audio segments (e.g. 'en-US' for English, 'zh-CN' for Chinese, 'ja-JP' for Japanese, etc.). 
You will also need to provide the [path] for both the audio segments and the runsheet. 

Complete language code list for Azure Speech Service: 
https://learn.microsoft.com/en-us/azure/cognitive-services/speech-service/language-support?tabs=stt

    """)
    try:
        language = input(">>> [Input 1/2] Please enter the language code (for example, 'en-US' for English): ")
        directory  = input(">>> [Input 2/2] Please enter the folder path for the audio segments and the runsheet: ")
    except Exception as e:
        print(">>> Error: " + str(e))
        input(">>> Press Enter or any key to exit..")
        return
    print(">>> [Transcribing]: Please be patient until finished and don't touch anything related...")
    
    try:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(['Filename', 'Script', 'Audio_exists', 'Transcript', 'Similarity'])
    except Exception as e:
        print(">>> Error: " + str(e))
        input(">>> Press Enter or any key to close me, then retry...")
        return
    
    try:
        # Find all audio files and the runsheet file in the directory
        audio_files = []
        runsheet_path = None
        for filename in os.listdir(directory):
            if filename.endswith(".wav") or filename.endswith(".mp3"):
                audio_files.append(filename)
            elif filename.endswith(".xlsx") and filename != "comparison_results.xlsx":
                runsheet_path = os.path.join(directory, filename)
    except Exception as e:
        print(">>> Error: " + str(e))
        input(">>> Press Enter or any key to close me, then retry...")
        return
    
    # If no runsheet file was found, terminate the program
    if runsheet_path is None:
        print(">>> Error: No runsheet file (.xlsx) found in the specified directory.")
        return
    else:
        try:
            runsheet = read_runsheet(runsheet_path)
        except Exception as e:
            print(">>> Error: " + str(e))
            input(">>> Press Enter or any key to close me, then retry...")
            return
    
    try:
        for filename, script in runsheet.items():  
            result = {'filename': filename, 'script': script, 'audio_exists': False, 'transcript': '', 'similarity': ''}
            if filename in audio_files:  
                audio_file_path = os.path.join(directory, filename)
                result['audio_exists'] = True
                transcript = transcribe_audio(audio_file_path, language)
                result['transcript'] = transcript
                similarity = compare_transcript_to_runsheet(transcript, script)
                result['similarity'] = format(similarity * 100, '.2f') + "%"  # convert to percentage and format with 2 decimal places
            sheet.append([result['filename'], result['script'], result['audio_exists'], result['transcript'], result['similarity']])
    except Exception as e:
        print(">>> Error: " + str(e))
        input(">>> Press Enter or any key to close me, then retry...")
        return
    try:
        wb.save(directory + 'comparison_results.xlsx')
    except Exception as e:
        print(">>> Error: " + str(e))
        input(">>> Press Enter or any key to close me, then retry...")
        return
    input(">>> [Completed] Results have been saved to 'comparison_results.xlsx'. Press Enter or close this window..")

if __name__ == "__main__":
    main()